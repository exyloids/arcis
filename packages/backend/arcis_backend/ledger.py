"""Manual-ledger application services backed by PostgreSQL.

This is deliberately small but keeps the important boundary: uploaded rows are
staged first, then confirmation atomically creates immutable source records and
canonical transactions.
"""

from __future__ import annotations

import base64
import binascii
import calendar
import csv
import hashlib
import io
import json
import re
from collections import defaultdict
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from statistics import median
from uuid import UUID, uuid4

from openpyxl import load_workbook
from sqlalchemy import Engine, create_engine, text
from sqlalchemy.orm import Session

from arcis_backend.storage import MinioArtifactStorage

DEFAULT_CATEGORIES = (
    ("food_dining", "Food and Dining"),
    ("groceries", "Groceries"),
    ("shopping", "Shopping"),
    ("bills_utilities", "Bills and Utilities"),
    ("travel", "Travel"),
    ("transportation", "Transportation"),
    ("salary_income", "Salary and Income"),
    ("transfers", "Transfers"),
    ("cash_withdrawal", "Cash Withdrawal"),
    ("fees_charges", "Fees and Charges"),
    ("other", "Other"),
)

CATEGORY_TAXONOMY = {
    "transport": ("Transport", "Uber, Rapido, Auto, Cab, Train, Metro, Bus, Bike, Fuel, EV Recharge, Flights, Parking, FastTag, Tolls, Lounge, Fine"),
    "food_drinks": ("Food & Drinks", "Eating Out, Take Away, Tea & Coffee, Fast Food, Snacks, Swiggy, Zomato, Sweets, Liquor, Beverages, Date, Pizza, Tiffin"),
    "shopping": ("Shopping", "Clothes, Footwear, Electronics, Festival, Video Games, Books, Plants, Jewellery, Furniture, Appliances, Utensils, Vehicle, Cosmetics, Toys, Stationery"),
    "groceries": ("Groceries", "Supermarket, Fruits & Vegetables, Dairy, Meat & Seafood, Household Supplies"),
    "home": ("Home", "Rent, Maintenance, Repairs, Furnishing, Domestic Help"),
    "entertainment": ("Entertainment", "Movies, Streaming, Games, Music, Hobbies"),
    "events": ("Events", "Tickets, Celebrations, Gifts, Conferences"),
    "travel": ("Travel", "Hotels, Bookings, Visa, Foreign Exchange"),
    "medical": ("Medical", "Doctor, Pharmacy, Tests, Hospital, Insurance"),
    "personal": ("Personal", "Salon, Clothing Care, Mobile, Miscellaneous"),
    "fitness": ("Fitness", "Gym, Sports, Wellness"),
    "services": ("Services", "Professional, Repairs, Delivery, Government"),
    "bills": ("Bills", "Electricity, Water, Internet, Mobile, Gas"),
    "subscriptions": ("Subscriptions", "Software, Streaming, Memberships"),
    "emi": ("EMI", "Home Loan, Vehicle Loan, Personal Loan"),
    "credit_bill": ("Credit Bill", "Credit Card Bill Payment"),
}

BUILTIN_MERCHANT_MAPPINGS = (
    ("swiggy", "Swiggy", "food_drinks_swiggy"),
    ("zomato", "Zomato", "food_drinks_zomato"),
    ("uber", "Uber", "transport_uber"),
    ("rapido", "Rapido", "transport_rapido"),
    ("makemytrip", "MakeMyTrip", "transport_flights"),
    ("indianoil", "IndianOil", "transport_fuel"),
    ("hpcl", "HP Fuel", "transport_fuel"),
    ("bpcl", "BP Fuel", "transport_fuel"),
    ("amazon", "Amazon", "shopping_electronics"),
)

REPORTING_PERIODS = {
    "all_time",
    "this_month",
    "last_month",
    "last_3_months",
    "last_6_months",
    "this_year",
}


class LedgerError(ValueError):
    """A safe, user-correctable ledger operation error."""


def database_engine(database_url: str) -> Engine:
    return create_engine(database_url, pool_pre_ping=True)


class LedgerService:
    def __init__(self, engine: Engine, user_id: UUID, storage: MinioArtifactStorage | None = None) -> None:
        self.engine = engine
        self.user_id = user_id
        self.storage = storage

    def initialize_user(self) -> None:
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    """
                    INSERT INTO users (id, email_normalized, display_name)
                    VALUES (:id, :email, 'Local Arcis User')
                    ON CONFLICT (id) DO NOTHING
                    """
                ),
                {"id": self.user_id, "email": f"local-{self.user_id}@arcis.invalid"},
            )
            for code, name in DEFAULT_CATEGORIES:
                session.execute(
                    text(
                        """
                        INSERT INTO categories (id, user_id, code, name, is_system)
                        VALUES (:id, :user_id, :code, :name, true)
                        ON CONFLICT (user_id, code) DO NOTHING
                        """
                    ),
                    {"id": uuid4(), "user_id": self.user_id, "code": code, "name": name},
                )
            for code, (name, children) in CATEGORY_TAXONOMY.items():
                parent_id = uuid4()
                session.execute(text("""INSERT INTO categories (id, user_id, code, name, is_system)
                    VALUES (:id, :user_id, :code, :name, true) ON CONFLICT (user_id, code) DO NOTHING"""),
                    {"id": parent_id, "user_id": self.user_id, "code": code, "name": name})
                parent_id = session.execute(text("SELECT id FROM categories WHERE user_id = :user_id AND code = :code"), {"user_id": self.user_id, "code": code}).scalar_one()
                for child in children.split(", "):
                    child_code = f"{code}_{child.lower().replace(' & ', '_').replace(' ', '_')}"
                    session.execute(text("""INSERT INTO categories (id, user_id, code, name, parent_id, is_system)
                        VALUES (:id, :user_id, :code, :name, :parent_id, true) ON CONFLICT (user_id, code) DO NOTHING"""),
                        {"id": uuid4(), "user_id": self.user_id, "code": child_code, "name": child, "parent_id": parent_id})
            self._seed_builtin_merchant_rules(session)

    def _seed_builtin_merchant_rules(self, session: Session) -> None:
        """Register application-maintained keyword rules once per user."""
        for pattern, merchant, category_code in BUILTIN_MERCHANT_MAPPINGS:
            category_id = session.execute(
                text("SELECT id FROM categories WHERE user_id = :user_id AND code = :code"),
                {"user_id": self.user_id, "code": category_code},
            ).scalar_one_or_none()
            if category_id is None:
                continue
            session.execute(
                text(
                    """INSERT INTO merchant_rules (id, user_id, match_pattern, merchant_normalized,
                        category_id, priority, rule_type, confidence)
                    VALUES (:id, :user_id, :pattern, :merchant, :category_id, 100, 'keyword', 0.9500)
                    ON CONFLICT (user_id, match_pattern) DO NOTHING"""
                ),
                {"id": uuid4(), "user_id": self.user_id, "pattern": pattern, "merchant": merchant,
                 "category_id": category_id},
            )

    def list_accounts(self) -> list[dict[str, object]]:
        return self._rows(
            """
            SELECT id, account_type, institution_code, product_name, display_name,
                   masked_identifier, currency, status, version
            FROM financial_accounts WHERE user_id = :user_id AND status = 'active'
            ORDER BY display_name
            """
        )

    def create_account(self, payload: dict[str, object]) -> dict[str, object]:
        account_id = uuid4()
        account_type = _required_choice(payload, "account_type", {"bank_account", "credit_card"})
        display_name = _required_text(payload, "display_name")
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    """
                    INSERT INTO financial_accounts
                    (id, user_id, account_type, institution_code, product_name, display_name,
                     masked_identifier, currency)
                    VALUES (:id, :user_id, :account_type, :institution_code, :product_name,
                            :display_name, :masked_identifier, :currency)
                    """
                ),
                {
                    "id": account_id,
                    "user_id": self.user_id,
                    "account_type": account_type,
                    "institution_code": _required_text(payload, "institution_code").lower(),
                    "product_name": _required_text(payload, "product_name"),
                    "display_name": display_name,
                    "masked_identifier": _optional_text(payload, "masked_identifier"),
                    "currency": _optional_text(payload, "currency") or "INR",
                },
            )
        return self._one("SELECT * FROM financial_accounts WHERE id = :id", {"id": account_id})

    def list_categories(self) -> list[dict[str, object]]:
        return self._rows(
            """SELECT c.id, c.code, c.name, c.parent_id, p.name AS parent_name, c.is_system, c.version
            FROM categories c LEFT JOIN categories p ON p.id = c.parent_id
            WHERE c.user_id = :user_id AND c.archived_at IS NULL
            ORDER BY COALESCE(p.name, c.name), c.parent_id NULLS FIRST, c.name"""
        )

    def create_category(self, payload: dict[str, object]) -> dict[str, object]:
        category_id = uuid4()
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    "INSERT INTO categories (id, user_id, code, name) VALUES (:id, :user_id, :code, :name)"
                ),
                {
                    "id": category_id,
                    "user_id": self.user_id,
                    "code": _required_text(payload, "code").lower().replace(" ", "_"),
                    "name": _required_text(payload, "name"),
                },
            )
        return self._one("SELECT * FROM categories WHERE id = :id", {"id": category_id})

    def list_merchant_rules(self) -> list[dict[str, object]]:
        return self._rows("""SELECT mr.id, mr.match_pattern, mr.merchant_normalized, mr.priority,
            mr.category_id, c.name AS category FROM merchant_rules mr LEFT JOIN categories c ON c.id = mr.category_id
            WHERE mr.user_id = :user_id ORDER BY mr.priority, mr.created_at""")

    def create_merchant_rule(self, payload: dict[str, object]) -> dict[str, object]:
        rule_id = uuid4()
        category_id = UUID(str(payload["category_id"])) if payload.get("category_id") else None
        with Session(self.engine) as session, session.begin():
            session.execute(text("""INSERT INTO merchant_rules (id, user_id, match_pattern, merchant_normalized, category_id, priority)
                VALUES (:id, :user_id, :pattern, :merchant, :category_id, :priority)"""),
                {"id": rule_id, "user_id": self.user_id, "pattern": _required_text(payload, "match_pattern").upper(),
                 "merchant": _required_text(payload, "merchant_normalized"), "category_id": category_id,
                 "priority": int(payload.get("priority", 100))})
        return self._one("SELECT * FROM merchant_rules WHERE id = :id", {"id": rule_id})

    def apply_merchant_rules(self) -> dict[str, int]:
        return self.categorize_transactions()

    def apply_builtin_categories(self) -> dict[str, int]:
        return self.categorize_transactions()

    def categorize_transactions(self) -> dict[str, int]:
        """Apply deterministic rules: override, exact merchant, MCC, then keyword."""
        with Session(self.engine) as session, session.begin():
            rules = session.execute(text("""SELECT * FROM merchant_rules WHERE user_id = :user_id
                ORDER BY CASE rule_type WHEN 'user_override' THEN 1 WHEN 'exact_merchant' THEN 2 WHEN 'mcc' THEN 3 ELSE 4 END,
                priority, created_at"""), {"user_id": self.user_id}).mappings().all()
            transactions = session.execute(text("""SELECT id, narration, merchant_normalized, merchant_mcc,
                category_source FROM transactions WHERE user_id = :user_id"""), {"user_id": self.user_id}).mappings().all()
            updated = 0
            for transaction in transactions:
                if transaction["category_source"] == "manual":
                    continue
                merchant = _normalize_merchant(transaction["merchant_normalized"] or transaction["narration"])
                for rule in rules:
                    normalized_pattern = _normalize_merchant(rule["match_pattern"])
                    matched = (
                        rule["rule_type"] == "mcc" and transaction["merchant_mcc"] == rule["match_pattern"]
                    ) or (
                        rule["rule_type"] in {"user_override", "exact_merchant"}
                        and merchant == normalized_pattern
                    ) or (
                        rule["rule_type"] == "keyword" and normalized_pattern in merchant
                    )
                    if not matched:
                        continue
                    session.execute(text("""UPDATE transactions SET merchant_normalized = :merchant, category_id = :category_id,
                        category_source = :source, category_rule_id = :rule_id, category_confidence = :confidence, updated_at = now()
                        WHERE id = :id"""), {"merchant": rule["merchant_normalized"], "category_id": rule["category_id"], "source": rule["rule_type"], "rule_id": rule["id"], "confidence": rule["confidence"], "id": transaction["id"]})
                    updated += 1
                    break
        return {"rules": len(rules), "transactions_updated": updated}

    def stage_import(
        self, account_id: UUID, filename: str, content: bytes, column_mapping: dict[str, str] | None = None
    ) -> dict[str, object]:
        self._require_account(account_id)
        rows, row_errors = _parse_tabular_upload_with_issues(filename, content, column_mapping)
        if not rows and not row_errors:
            raise LedgerError("The uploaded statement has no transaction rows")
        import_id = uuid4()
        content_hash = hashlib.sha256(content).hexdigest()
        with Session(self.engine) as session:
            existing_import_id = session.execute(
                text(
                    """SELECT id FROM imports WHERE user_id = :user_id
                    AND financial_account_id = :account_id AND content_sha256 = :content_sha256"""
                ),
                {"user_id": self.user_id, "account_id": account_id, "content_sha256": content_hash},
            ).scalar_one_or_none()
        if existing_import_id is not None:
            return self.import_preview(existing_import_id)
        stored = None
        if self.storage is not None:
            stored = self.storage.put(self.user_id, import_id, _safe_filename(filename), content)
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    """
                    INSERT INTO imports (id, user_id, financial_account_id, filename, content_sha256, state,
                                         row_count, valid_row_count, invalid_row_count, error_code, object_key,
                                         detected_mime_type, byte_size)
                    VALUES (:id, :user_id, :account_id, :filename, :content_sha256, :state, :row_count,
                            :valid_row_count, :invalid_row_count, :error_code, :object_key,
                            :detected_mime_type, :byte_size)
                    """
                ),
                {
                    "id": import_id,
                    "user_id": self.user_id,
                    "account_id": account_id,
                    "filename": _safe_filename(filename),
                    "content_sha256": content_hash,
                    "state": "preview_ready" if rows else "failed",
                    "row_count": len(rows) + len(row_errors),
                    "valid_row_count": len(rows),
                    "invalid_row_count": len(row_errors),
                    "error_code": "row_validation_failed" if row_errors else None,
                    "object_key": stored.object_key if stored else None,
                    "detected_mime_type": stored.content_type if stored else None,
                    "byte_size": stored.byte_size if stored else len(content),
                },
            )
            for ordinal, row in enumerate(rows, start=1):
                session.execute(
                    text(
                        """
                        INSERT INTO import_rows
                        (id, import_id, ordinal, transaction_date, posted_date, narration, amount, currency,
                         direction, provider_reference, raw_columns)
                        VALUES (:id, :import_id, :ordinal, :transaction_date, :posted_date, :narration,
                                :amount, :currency, :direction, :provider_reference, CAST(:raw_columns AS jsonb))
                        """
                    ),
                    {"id": uuid4(), "import_id": import_id, "ordinal": ordinal, **row},
                )
            for error in row_errors:
                session.execute(
                    text(
                        """INSERT INTO import_row_errors (id, import_id, ordinal, message)
                        VALUES (:id, :import_id, :ordinal, :message)"""
                    ),
                    {"id": uuid4(), "import_id": import_id, **error},
                )
        return self.import_preview(import_id)

    def import_preview(self, import_id: UUID) -> dict[str, object]:
        import_row = self._one(
            "SELECT id, financial_account_id, filename, state, row_count, valid_row_count, invalid_row_count, "
            "created_at FROM imports "
            "WHERE id = :id AND user_id = :user_id",
            {"id": import_id, "user_id": self.user_id},
        )
        rows = self._rows(
            "SELECT ordinal, transaction_date, posted_date, narration, amount, currency, direction, "
            "provider_reference FROM import_rows WHERE import_id = :id ORDER BY ordinal",
            {"id": import_id},
        )
        errors = self._rows(
            "SELECT ordinal, message FROM import_row_errors WHERE import_id = :id ORDER BY ordinal",
            {"id": import_id},
        )
        return {"import": import_row, "rows": rows, "errors": errors}

    def list_imports(self) -> list[dict[str, object]]:
        return self._rows(
            """SELECT id, financial_account_id, filename, state, row_count, valid_row_count, invalid_row_count,
            duplicate_count,
            confirmed_at, created_at FROM imports WHERE user_id = :user_id ORDER BY created_at DESC"""
        )

    def list_documents(self) -> list[dict[str, object]]:
        """Return safe document metadata without exposing object keys or contents."""
        return self._rows(
            """SELECT sa.id, sa.kind, sa.detected_mime_type, sa.byte_size,
                      sa.lifecycle_state, sa.created_at, sa.deleted_at, sa.purge_after,
                      (sa.recovery_object_key IS NOT NULL) AS can_restore,
                      i.id AS import_id, i.filename, i.state AS import_state,
                      a.display_name AS account_name,
                      m.display_email AS mailbox_email,
                      pc.state AS parser_state, pc.review_reason
               FROM source_artifacts sa
               LEFT JOIN imports i ON i.id = sa.import_id
               LEFT JOIN financial_accounts a ON a.id = i.financial_account_id
               LEFT JOIN mailboxes m ON m.id = sa.mailbox_id
               LEFT JOIN LATERAL (
                   SELECT state, review_reason
                   FROM parser_candidates
                   WHERE artifact_id = sa.id
                   ORDER BY created_at DESC
                   LIMIT 1
               ) pc ON true
               WHERE sa.user_id = :user_id
               ORDER BY sa.created_at DESC"""
        )

    def redact_document(self, artifact_id: UUID) -> dict[str, object]:
        """Move raw content to a bounded recovery area while retaining provenance."""
        with Session(self.engine) as session, session.begin():
            artifact = session.execute(
                text(
                    """SELECT id, object_key, recovery_object_key, lifecycle_state
                       FROM source_artifacts
                       WHERE id = :id AND user_id = :user_id
                       FOR UPDATE"""
                ),
                {"id": artifact_id, "user_id": self.user_id},
            ).mappings().one_or_none()
            if artifact is None:
                raise LedgerError("Document was not found")
            if artifact["lifecycle_state"] == "redacted":
                return {
                    "id": artifact_id,
                    "lifecycle_state": "redacted",
                    "can_restore": artifact["recovery_object_key"] is not None,
                }
            object_key = artifact["object_key"]
            recovery_key = None
            if object_key:
                if self.storage is None:
                    raise LedgerError("Document storage is unavailable")
                recovery_key = self.storage.quarantine(
                    str(object_key), self.user_id, artifact_id
                )
            session.execute(
                text(
                    """UPDATE source_artifacts
                       SET original_object_key = object_key,
                           recovery_object_key = :recovery_object_key,
                           object_key = NULL,
                           lifecycle_state = 'redacted',
                           deleted_at = now(),
                           purge_after = now() + INTERVAL '30 days'
                       WHERE id = :id"""
                ),
                {"id": artifact_id, "recovery_object_key": recovery_key},
            )
            session.execute(
                text(
                    """INSERT INTO audit_events
                       (id, user_id, actor_type, action, target_type, target_id, result)
                       VALUES (:id, :user_id, 'user', 'document.redact',
                               'source_artifact', :target_id, 'success')"""
                ),
                {
                    "id": uuid4(),
                    "user_id": self.user_id,
                    "target_id": artifact_id,
                },
            )
        return {
            "id": artifact_id,
            "lifecycle_state": "redacted",
            "can_restore": recovery_key is not None,
        }

    def restore_document(self, artifact_id: UUID) -> dict[str, object]:
        with Session(self.engine) as session, session.begin():
            artifact = session.execute(
                text(
                    """SELECT id, original_object_key, recovery_object_key,
                              lifecycle_state
                       FROM source_artifacts
                       WHERE id = :id AND user_id = :user_id
                       FOR UPDATE"""
                ),
                {"id": artifact_id, "user_id": self.user_id},
            ).mappings().one_or_none()
            if artifact is None:
                raise LedgerError("Document was not found")
            if artifact["lifecycle_state"] != "redacted":
                raise LedgerError("Only a deleted document can be restored")
            original_key = artifact["original_object_key"]
            recovery_key = artifact["recovery_object_key"]
            if not original_key or not recovery_key or self.storage is None:
                raise LedgerError("This document no longer has recoverable content")
            self.storage.restore(str(recovery_key), str(original_key))
            session.execute(
                text(
                    """UPDATE source_artifacts
                       SET object_key = original_object_key,
                           original_object_key = NULL,
                           recovery_object_key = NULL,
                           lifecycle_state = 'active',
                           deleted_at = NULL,
                           purge_after = NULL
                       WHERE id = :id"""
                ),
                {"id": artifact_id},
            )
            session.execute(
                text(
                    """INSERT INTO audit_events
                       (id, user_id, actor_type, action, target_type, target_id, result)
                       VALUES (:id, :user_id, 'user', 'document.restore',
                               'source_artifact', :target_id, 'success')"""
                ),
                {
                    "id": uuid4(),
                    "user_id": self.user_id,
                    "target_id": artifact_id,
                },
            )
        return {"id": artifact_id, "lifecycle_state": "active"}

    def cancel_import(self, import_id: UUID) -> None:
        with Session(self.engine) as session, session.begin():
            imported = session.execute(
                text("SELECT state, object_key FROM imports WHERE id = :id AND user_id = :user_id FOR UPDATE"),
                {"id": import_id, "user_id": self.user_id},
            ).mappings().one_or_none()
            if imported is None:
                raise LedgerError("Import was not found")
            if imported["state"] == "confirmed":
                raise LedgerError("Confirmed imports cannot be cancelled")
            session.execute(
                text("UPDATE imports SET state = 'cancelled', cancelled_at = now() WHERE id = :id"),
                {"id": import_id},
            )
        if imported["object_key"] and self.storage is not None:
            self.storage.delete(imported["object_key"])

    def confirm_import(self, import_id: UUID) -> dict[str, int]:
        with Session(self.engine) as session, session.begin():
            imported = session.execute(
                text(
                    "SELECT * FROM imports WHERE id = :id AND user_id = :user_id FOR UPDATE"
                ),
                {"id": import_id, "user_id": self.user_id},
            ).mappings().one_or_none()
            if imported is None:
                raise LedgerError("Import was not found")
            if imported["state"] == "confirmed":
                return {"created": 0, "duplicates": imported["duplicate_count"], "confirmed": 1}
            if imported["state"] != "preview_ready":
                raise LedgerError(f"Import cannot be confirmed from state {imported['state']}")
            artifact_id = uuid4()
            session.execute(
                text(
                    """
                    INSERT INTO source_artifacts (id, user_id, kind, content_sha256, detected_mime_type,
                                                  lifecycle_state, import_id, object_key, byte_size)
                    VALUES (:id, :user_id, 'manual_upload', :content_sha256, 'text/tabular', 'active', :import_id,
                            :object_key, :byte_size)
                    ON CONFLICT (user_id, kind, content_sha256) DO NOTHING
                    """
                ),
                {"id": artifact_id, "user_id": self.user_id, "content_sha256": imported["content_sha256"], "import_id": import_id, "object_key": imported["object_key"], "byte_size": imported["byte_size"]},
            )
            artifact = session.execute(
                text("SELECT id FROM source_artifacts WHERE user_id = :user_id AND kind = 'manual_upload' "
                     "AND content_sha256 = :content_sha256"),
                {"user_id": self.user_id, "content_sha256": imported["content_sha256"]},
            ).scalar_one()
            created = duplicates = 0
            for row in session.execute(
                text("SELECT * FROM import_rows WHERE import_id = :id ORDER BY ordinal"), {"id": import_id}
            ).mappings():
                source_id = uuid4()
                source_key = f"{import_id}:{row['ordinal']}"
                source_insert = session.execute(
                    text(
                        """
                        INSERT INTO source_records (id, user_id, artifact_id, source_record_key, transaction_date,
                                                    posted_date, narration, amount, currency, direction, provider_reference)
                        VALUES (:id, :user_id, :artifact_id, :source_record_key, :transaction_date, :posted_date,
                                :narration, :amount, :currency, :direction, :provider_reference)
                        ON CONFLICT (artifact_id, source_record_key) DO NOTHING
                        """
                    ),
                    {"id": source_id, "user_id": self.user_id, "artifact_id": artifact, "source_record_key": source_key,
                     "transaction_date": row["transaction_date"], "posted_date": row["posted_date"],
                     "narration": row["narration"], "amount": row["amount"], "currency": row["currency"],
                     "direction": row["direction"], "provider_reference": row["provider_reference"]},
                )
                if source_insert.rowcount == 0:
                    duplicates += 1
                    continue
                duplicate = session.execute(
                    text(
                        """SELECT id FROM transactions WHERE user_id = :user_id AND financial_account_id = :account_id
                        AND transaction_date = :transaction_date AND amount = :amount AND direction = :direction
                        AND COALESCE(provider_reference, '') = COALESCE(:provider_reference, '') LIMIT 1"""
                    ),
                    {"user_id": self.user_id, "account_id": imported["financial_account_id"],
                     "transaction_date": row["transaction_date"], "amount": row["amount"], "direction": row["direction"],
                     "provider_reference": row["provider_reference"]},
                ).scalar_one_or_none()
                if duplicate is not None:
                    duplicates += 1
                    continue
                kind = _transaction_kind(row["narration"])
                transaction_id = uuid4()
                session.execute(
                    text(
                        """INSERT INTO transactions (id, user_id, financial_account_id, transaction_date, posted_date,
                                                     narration, amount, currency, direction, transaction_kind,
                                                     reconciliation_state, source_record_id)
                        VALUES (:id, :user_id, :account_id, :transaction_date, :posted_date, :narration, :amount,
                                :currency, :direction, :transaction_kind, 'statement_confirmed', :source_record_id)"""
                    ),
                    {"id": transaction_id, "user_id": self.user_id, "account_id": imported["financial_account_id"],
                     "transaction_date": row["transaction_date"], "posted_date": row["posted_date"],
                     "narration": row["narration"], "amount": row["amount"], "currency": row["currency"],
                     "direction": row["direction"], "transaction_kind": kind, "source_record_id": source_id},
                )
                session.execute(
                    text("INSERT INTO transaction_evidence (transaction_id, source_record_id, relationship, match_method) "
                         "VALUES (:transaction_id, :source_record_id, 'primary', 'import_confirmation')"),
                    {"transaction_id": transaction_id, "source_record_id": source_id},
                )
                created += 1
            session.execute(
                text("UPDATE imports SET state = 'confirmed', confirmed_at = now(), duplicate_count = :duplicates "
                     "WHERE id = :id"), {"duplicates": duplicates, "id": import_id}
            )
        return {"created": created, "duplicates": duplicates, "confirmed": 1}

    def list_transactions(
        self,
        *,
        month: str | None = None,
        period: str | None = None,
        account_id: UUID | None = None,
        account_type: str | None = None,
        category_id: UUID | None = None,
        query_text: str | None = None,
        limit: int = 100,
    ) -> list[dict[str, object]]:
        query = """
            SELECT t.id, t.financial_account_id, a.display_name AS account_name, t.transaction_date,
                   t.posted_date, t.narration, t.merchant_normalized, t.provider_reference, t.amount, t.currency, t.direction, t.transaction_kind,
                   t.reconciliation_state, t.category_id, COALESCE(c.name, CASE WHEN t.transaction_kind = 'credit_card_payment' THEN 'Credit Card Bill Payment' END) AS category
            FROM transactions t JOIN financial_accounts a ON a.id = t.financial_account_id
            LEFT JOIN categories c ON c.id = t.category_id
            WHERE t.user_id = :user_id
        """
        parameters: dict[str, object] = {"user_id": self.user_id}
        if month:
            query += " AND to_char(t.transaction_date, 'YYYY-MM') = :month"
            parameters["month"] = month
        elif period:
            query, parameters = _apply_period_filter(query, parameters, "t.transaction_date", period)
        if account_id:
            query += " AND t.financial_account_id = :account_id"
            parameters["account_id"] = account_id
        if account_type:
            query += " AND a.account_type = :account_type"
            parameters["account_type"] = account_type
        if category_id:
            query += " AND t.category_id = :category_id"
            parameters["category_id"] = category_id
        if query_text:
            query += " AND t.narration ILIKE :query_text"
            parameters["query_text"] = f"%{query_text.strip()}%"
        parameters["limit"] = min(max(limit, 1), 200)
        return self._rows(query + " ORDER BY t.transaction_date DESC, t.created_at DESC LIMIT :limit", parameters)

    def transaction_page(
        self,
        *,
        month: str | None = None,
        period: str | None = None,
        account_id: UUID | None = None,
        account_type: str | None = None,
        category_id: UUID | None = None,
        query_text: str | None = None,
        cursor: str | None = None,
        limit: int = 50,
    ) -> dict[str, object]:
        query = """
            SELECT t.id, t.financial_account_id, a.display_name AS account_name, t.transaction_date,
                   t.posted_date, t.narration, t.merchant_normalized, t.provider_reference, t.amount, t.currency, t.direction, t.transaction_kind,
                   t.reconciliation_state, t.category_id, COALESCE(c.name, CASE WHEN t.transaction_kind = 'credit_card_payment' THEN 'Credit Card Bill Payment' END) AS category
            FROM transactions t JOIN financial_accounts a ON a.id = t.financial_account_id
            LEFT JOIN categories c ON c.id = t.category_id
            WHERE t.user_id = :user_id
        """
        parameters: dict[str, object] = {"user_id": self.user_id}
        if month:
            query += " AND to_char(t.transaction_date, 'YYYY-MM') = :month"
            parameters["month"] = month
        elif period:
            query, parameters = _apply_period_filter(query, parameters, "t.transaction_date", period)
        if account_id:
            query += " AND t.financial_account_id = :account_id"
            parameters["account_id"] = account_id
        if account_type:
            query += " AND a.account_type = :account_type"
            parameters["account_type"] = account_type
        if category_id:
            query += " AND t.category_id = :category_id"
            parameters["category_id"] = category_id
        if query_text:
            query += " AND t.narration ILIKE :query_text"
            parameters["query_text"] = f"%{query_text.strip()}%"
        if cursor:
            cursor_date, cursor_id = _decode_transaction_cursor(cursor)
            query += " AND (t.transaction_date, t.id) < (:cursor_date, :cursor_id)"
            parameters.update({"cursor_date": cursor_date, "cursor_id": cursor_id})
        requested_limit = min(max(limit, 1), 200)
        parameters["limit"] = requested_limit + 1
        rows = self._rows(query + " ORDER BY t.transaction_date DESC, t.id DESC LIMIT :limit", parameters)
        has_more = len(rows) > requested_limit
        items = rows[:requested_limit]
        next_cursor = None
        if has_more and items:
            next_cursor = _encode_transaction_cursor(items[-1]["transaction_date"], items[-1]["id"])
        return {"items": items, "next_cursor": next_cursor}

    def transaction_evidence(self, transaction_id: UUID) -> list[dict[str, object]]:
        return self._rows(
            """SELECT sr.id AS source_record_id, sr.source_record_key, sr.transaction_date, sr.narration,
            sr.amount, sr.currency, sr.direction, sr.provider_reference, sa.kind AS artifact_kind,
            i.id AS import_id, i.filename AS import_filename
            FROM transaction_evidence te JOIN source_records sr ON sr.id = te.source_record_id
            JOIN source_artifacts sa ON sa.id = sr.artifact_id
            LEFT JOIN imports i ON i.id = sa.import_id
            WHERE te.transaction_id = :transaction_id AND sr.user_id = :user_id""",
            {"transaction_id": transaction_id, "user_id": self.user_id},
        )

    def update_transaction(self, transaction_id: UUID, payload: dict[str, object]) -> dict[str, object]:
        fields: list[str] = []
        parameters: dict[str, object] = {"id": transaction_id, "user_id": self.user_id}
        for field in ("narration", "transaction_kind"):
            if field in payload:
                fields.append(f"{field} = :{field}")
                parameters[field] = _required_text(payload, field)
        if "category_id" in payload:
            fields.extend(("category_id = :category_id", "category_source = 'manual'", "category_rule_id = NULL", "category_confidence = 1.0"))
            parameters["category_id"] = UUID(str(payload["category_id"])) if payload["category_id"] else None
        if not fields:
            raise LedgerError("No mutable transaction fields were supplied")
        fields.append("version = version + 1")
        fields.append("updated_at = now()")
        with Session(self.engine) as session, session.begin():
            result = session.execute(
                text("UPDATE transactions SET " + ", ".join(fields) + " WHERE id = :id AND user_id = :user_id"),
                parameters,
            )
            if result.rowcount != 1:
                raise LedgerError("Transaction was not found")
            if payload.get("remember_merchant") and payload.get("category_id"):
                transaction = session.execute(text("SELECT narration, merchant_normalized FROM transactions WHERE id = :id AND user_id = :user_id"), parameters).mappings().one()
                merchant = _normalize_merchant(transaction["merchant_normalized"] or transaction["narration"])
                if merchant:
                    session.execute(text("""INSERT INTO merchant_rules (id, user_id, match_pattern, merchant_normalized,
                        category_id, priority, rule_type, confidence) VALUES (:id, :user_id, :pattern, :merchant,
                        :category_id, 1, 'user_override', 1.0) ON CONFLICT (user_id, match_pattern) DO UPDATE
                        SET category_id = EXCLUDED.category_id, merchant_normalized = EXCLUDED.merchant_normalized,
                        rule_type = 'user_override', confidence = 1.0, priority = 1, updated_at = now()"""),
                        {"id": uuid4(), "user_id": self.user_id, "pattern": merchant, "merchant": transaction["merchant_normalized"] or transaction["narration"], "category_id": UUID(str(payload["category_id"]))})
        return self._one("SELECT * FROM transactions WHERE id = :id", {"id": transaction_id})

    def monthly_report(self, month: str) -> dict[str, object]:
        rows = self._rows(
            """SELECT COALESCE(c.name, 'Uncategorized') AS category, t.direction, SUM(t.amount) AS amount
            FROM transactions t LEFT JOIN categories c ON c.id = t.category_id
            WHERE t.user_id = :user_id AND to_char(t.transaction_date, 'YYYY-MM') = :month
              AND t.transaction_kind NOT IN ('transfer', 'credit_card_payment')
            GROUP BY c.name, t.direction ORDER BY amount DESC""",
            {"user_id": self.user_id, "month": month},
        )
        totals = {"income": Decimal("0"), "expense": Decimal("0")}
        for row in rows:
            totals["expense" if row["direction"] == "debit" else "income"] += row["amount"]
        return {"month": month, "income": totals["income"], "expense": totals["expense"], "categories": rows}

    def preferences(self) -> dict[str, object]:
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    """INSERT INTO user_preferences (user_id)
                       VALUES (:user_id) ON CONFLICT (user_id) DO NOTHING"""
                ),
                {"user_id": self.user_id},
            )
            row = session.execute(
                text(
                    """SELECT reporting_period, retention_policy, updated_at
                       FROM user_preferences WHERE user_id = :user_id"""
                ),
                {"user_id": self.user_id},
            ).mappings().one()
            return dict(row)

    def update_reporting_period(self, period: str) -> dict[str, object]:
        if period not in REPORTING_PERIODS:
            raise LedgerError("Unsupported reporting period")
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    """INSERT INTO user_preferences (user_id, reporting_period)
                       VALUES (:user_id, :period)
                       ON CONFLICT (user_id) DO UPDATE
                       SET reporting_period = EXCLUDED.reporting_period, updated_at = now()"""
                ),
                {"user_id": self.user_id, "period": period},
            )
        return self.preferences()

    def update_retention_policy(self, payload: dict[str, object]) -> dict[str, object]:
        policy: dict[str, int] = {}
        limits = {
            "source_artifacts_days": (30, 3650),
            "statement_files_days": (30, 3650),
        }
        for field, (minimum, maximum) in limits.items():
            try:
                value = int(payload.get(field, 0))
            except (TypeError, ValueError) as error:
                raise LedgerError(f"{field} must be a whole number") from error
            if not minimum <= value <= maximum:
                raise LedgerError(f"{field} must be between {minimum} and {maximum} days")
            policy[field] = value
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    """INSERT INTO user_preferences (user_id, retention_policy)
                       VALUES (:user_id, CAST(:policy AS jsonb))
                       ON CONFLICT (user_id) DO UPDATE
                       SET retention_policy = EXCLUDED.retention_policy, updated_at = now()"""
                ),
                {"user_id": self.user_id, "policy": json.dumps(policy)},
            )
            session.execute(
                text(
                    """INSERT INTO audit_events
                       (id, user_id, actor_type, action, target_type, target_id, result,
                        safe_metadata)
                       VALUES (:id, :user_id, 'user', 'privacy.retention.update',
                               'user_preferences', :user_id, 'success', CAST(:policy AS jsonb))"""
                ),
                {
                    "id": uuid4(),
                    "user_id": self.user_id,
                    "policy": json.dumps(policy),
                },
            )
        return self.preferences()

    def enforce_retention_policy(self) -> dict[str, int]:
        policy = self.preferences()["retention_policy"]
        if not isinstance(policy, dict):
            raise LedgerError("Retention policy is invalid")
        now = datetime.now(UTC)
        source_cutoff = now - timedelta(days=int(policy["source_artifacts_days"]))
        statement_cutoff = now - timedelta(days=int(policy["statement_files_days"]))
        expired = self._rows(
            """SELECT id FROM source_artifacts
               WHERE user_id = :user_id AND lifecycle_state = 'active'
                 AND object_key IS NOT NULL
                 AND (
                   (kind = 'gmail_message' AND created_at < :source_cutoff)
                   OR
                   (kind IN ('manual_upload', 'gmail_attachment')
                    AND created_at < :statement_cutoff)
                 )
               ORDER BY created_at""",
            {
                "user_id": self.user_id,
                "source_cutoff": source_cutoff,
                "statement_cutoff": statement_cutoff,
            },
        )
        redacted = 0
        for artifact in expired:
            self.redact_document(UUID(str(artifact["id"])))
            redacted += 1

        purgeable = self._rows(
            """SELECT id, recovery_object_key FROM source_artifacts
               WHERE user_id = :user_id AND lifecycle_state = 'redacted'
                 AND recovery_object_key IS NOT NULL
                 AND purge_after <= now()""",
            {"user_id": self.user_id},
        )
        purged = 0
        for artifact in purgeable:
            if self.storage is None:
                raise LedgerError("Document storage is unavailable")
            self.storage.delete(str(artifact["recovery_object_key"]))
            with Session(self.engine) as session, session.begin():
                session.execute(
                    text(
                        """UPDATE source_artifacts
                           SET recovery_object_key = NULL,
                               original_object_key = NULL,
                               lifecycle_state = 'purged'
                           WHERE id = :id AND user_id = :user_id"""
                    ),
                    {"id": artifact["id"], "user_id": self.user_id},
                )
            purged += 1
        return {"redacted": redacted, "purged": purged}

    def privacy_inventory(self) -> dict[str, object]:
        counts = self._one(
            """SELECT
                   (SELECT COUNT(*) FROM financial_accounts WHERE user_id = :user_id) AS accounts,
                   (SELECT COUNT(*) FROM transactions WHERE user_id = :user_id) AS transactions,
                   (SELECT COUNT(*) FROM source_artifacts WHERE user_id = :user_id
                       AND lifecycle_state = 'active') AS stored_documents,
                   (SELECT COUNT(*) FROM mailboxes WHERE user_id = :user_id
                       AND connection_status = 'connected') AS connected_mailboxes""",
            {"user_id": self.user_id},
        )
        return {**counts, "retention_policy": self.preferences()["retention_policy"]}

    def privacy_export(self) -> dict[str, object]:
        """Create a portable export that deliberately excludes secrets and raw files."""
        user = self._one(
            """SELECT id, email_normalized, display_name, default_currency,
                      timezone, status, created_at, updated_at
               FROM users WHERE id = :user_id""",
            {"user_id": self.user_id},
        )
        collections = {
            "financial_accounts": self._rows(
                """SELECT id, account_type, institution_code, product_name, display_name,
                          masked_identifier, currency, status, created_at, updated_at
                   FROM financial_accounts WHERE user_id = :user_id ORDER BY created_at"""
            ),
            "transactions": self._rows(
                """SELECT id, financial_account_id, transaction_date, posted_date,
                          narration, amount, currency, direction, transaction_kind,
                          reconciliation_state, provider_reference, merchant_normalized,
                          category_id, category_source, created_at, updated_at
                   FROM transactions WHERE user_id = :user_id
                   ORDER BY transaction_date, id"""
            ),
            "categories": self._rows(
                """SELECT id, code, name, parent_id, is_system, created_at, updated_at
                   FROM categories WHERE user_id = :user_id ORDER BY name"""
            ),
            "budgets": self._rows(
                """SELECT id, category_id, monthly_limit, active, created_at, updated_at
                   FROM budgets WHERE user_id = :user_id ORDER BY created_at"""
            ),
            "recurring_payments": self._rows(
                """SELECT id, financial_account_id, display_name, category_id, cadence,
                          typical_amount, next_expected_on, confidence, state,
                          created_at, updated_at
                   FROM recurring_payment_detections
                   WHERE user_id = :user_id ORDER BY created_at"""
            ),
            "mailboxes": self._rows(
                """SELECT id, provider, display_email, connection_status, granted_scopes,
                          last_successful_sync_at, created_at, updated_at
                   FROM mailboxes WHERE user_id = :user_id ORDER BY created_at"""
            ),
            "documents": self.list_documents(),
        }
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    """INSERT INTO audit_events
                       (id, user_id, actor_type, action, target_type, target_id, result)
                       VALUES (:id, :user_id, 'user', 'privacy.export',
                               'user', :user_id, 'success')"""
                ),
                {"id": uuid4(), "user_id": self.user_id},
            )
        return {
            "schema_version": 1,
            "exported_on": date.today(),
            "user": user,
            **collections,
        }

    def period_report(self, period: str) -> dict[str, object]:
        if period not in REPORTING_PERIODS:
            raise LedgerError("Unsupported reporting period")
        query = """SELECT COALESCE(parent.name, c.name, 'Uncategorized') AS category,
                          t.direction, SUM(t.amount) AS amount
                   FROM transactions t
                   LEFT JOIN categories c ON c.id = t.category_id
                   LEFT JOIN categories parent ON parent.id = c.parent_id
                   WHERE t.user_id = :user_id
                     AND t.transaction_kind NOT IN ('transfer', 'credit_card_payment')"""
        parameters: dict[str, object] = {"user_id": self.user_id}
        query, parameters = _apply_period_filter(query, parameters, "t.transaction_date", period)
        rows = self._rows(
            query
            + """ GROUP BY COALESCE(parent.name, c.name, 'Uncategorized'), t.direction
                  ORDER BY amount DESC""",
            parameters,
        )
        totals = {"income": Decimal("0"), "expense": Decimal("0")}
        for row in rows:
            totals["expense" if row["direction"] == "debit" else "income"] += Decimal(str(row["amount"]))
        start, end = _reporting_period_bounds(period)
        return {
            "period": period,
            "date_from": start,
            "date_to": end - timedelta(days=1) if end else None,
            "income": totals["income"],
            "expense": totals["expense"],
            "categories": rows,
        }

    def list_budgets(self, month: str) -> list[dict[str, object]]:
        _month_start_end(month)
        rows = self._rows(
            """SELECT b.id, b.category_id, c.name AS category, b.monthly_limit,
                      b.active, COALESCE(spending.amount, 0) AS spent
               FROM budgets b
               JOIN categories c ON c.id = b.category_id
               LEFT JOIN (
                   SELECT COALESCE(category.parent_id, category.id) AS category_id,
                          SUM(t.amount) AS amount
                   FROM transactions t
                   LEFT JOIN categories category ON category.id = t.category_id
                   WHERE t.user_id = :user_id AND t.direction = 'debit'
                     AND to_char(t.transaction_date, 'YYYY-MM') = :month
                     AND t.transaction_kind NOT IN ('transfer', 'credit_card_payment')
                   GROUP BY COALESCE(category.parent_id, category.id)
               ) spending ON spending.category_id = b.category_id
               WHERE b.user_id = :user_id
               ORDER BY b.active DESC, c.name""",
            {"user_id": self.user_id, "month": month},
        )
        results = []
        for row in rows:
            limit = Decimal(str(row["monthly_limit"]))
            spent = Decimal(str(row["spent"]))
            results.append(
                {
                    **row,
                    "remaining": limit - spent,
                    "percentage": spent / limit * Decimal("100"),
                    "over_budget": spent > limit,
                }
            )
        return results

    def create_budget(self, payload: dict[str, object]) -> dict[str, object]:
        category_id = UUID(str(payload.get("category_id")))
        limit = _positive_decimal(payload.get("monthly_limit"), "monthly_limit")
        category = self._one(
            """SELECT id, parent_id FROM categories
               WHERE id = :id AND user_id = :user_id""",
            {"id": category_id, "user_id": self.user_id},
        )
        if category["parent_id"] is not None:
            raise LedgerError("Budgets must use a top-level category")
        budget_id = uuid4()
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    """INSERT INTO budgets
                       (id, user_id, category_id, monthly_limit, active)
                       VALUES (:id, :user_id, :category_id, :monthly_limit, true)
                       ON CONFLICT (user_id, category_id) DO UPDATE
                       SET monthly_limit = EXCLUDED.monthly_limit,
                           active = true, updated_at = now()"""
                ),
                {
                    "id": budget_id,
                    "user_id": self.user_id,
                    "category_id": category_id,
                    "monthly_limit": limit,
                },
            )
        return self._one(
            """SELECT b.id, b.category_id, c.name AS category, b.monthly_limit, b.active
               FROM budgets b JOIN categories c ON c.id = b.category_id
               WHERE b.user_id = :user_id AND b.category_id = :category_id""",
            {"user_id": self.user_id, "category_id": category_id},
        )

    def update_budget(self, budget_id: UUID, payload: dict[str, object]) -> dict[str, object]:
        fields: list[str] = []
        parameters: dict[str, object] = {"id": budget_id, "user_id": self.user_id}
        if "monthly_limit" in payload:
            fields.append("monthly_limit = :monthly_limit")
            parameters["monthly_limit"] = _positive_decimal(payload["monthly_limit"], "monthly_limit")
        if "active" in payload:
            fields.append("active = :active")
            parameters["active"] = bool(payload["active"])
        if not fields:
            raise LedgerError("No budget changes were provided")
        with Session(self.engine) as session, session.begin():
            result = session.execute(
                text(
                    f"""UPDATE budgets SET {", ".join(fields)}, updated_at = now()
                        WHERE id = :id AND user_id = :user_id"""
                ),
                parameters,
            )
            if result.rowcount != 1:
                raise LedgerError("Requested budget was not found")
        return self._one(
            """SELECT b.id, b.category_id, c.name AS category, b.monthly_limit, b.active
               FROM budgets b JOIN categories c ON c.id = b.category_id
               WHERE b.id = :id AND b.user_id = :user_id""",
            {"id": budget_id, "user_id": self.user_id},
        )

    def delete_budget(self, budget_id: UUID) -> None:
        with Session(self.engine) as session, session.begin():
            result = session.execute(
                text("DELETE FROM budgets WHERE id = :id AND user_id = :user_id"),
                {"id": budget_id, "user_id": self.user_id},
            )
            if result.rowcount != 1:
                raise LedgerError("Requested budget was not found")

    def spending_summary(self) -> dict[str, object]:
        """Return all-time expense categories grouped at the parent level."""
        rows = self._rows(
            """SELECT COALESCE(parent.id, c.id) AS category_id,
                       COALESCE(parent.name, c.name, 'Uncategorized') AS category,
                       SUM(t.amount) AS amount
                FROM transactions t
                LEFT JOIN categories c ON c.id = t.category_id
                LEFT JOIN categories parent ON parent.id = c.parent_id
                WHERE t.user_id = :user_id
                  AND t.direction = 'debit'
                  AND t.transaction_kind NOT IN ('transfer', 'credit_card_payment')
                GROUP BY COALESCE(parent.id, c.id), COALESCE(parent.name, c.name, 'Uncategorized')
                ORDER BY amount DESC""",
            {"user_id": self.user_id},
        )
        total = sum((Decimal(str(row["amount"])) for row in rows), Decimal("0"))
        categories = [
            {**row, "percentage": (Decimal(str(row["amount"])) / total * Decimal("100")) if total else Decimal("0")}
            for row in rows
        ]
        return {"expense": total, "categories": categories}

    def spending_category_trend(self, category_id: UUID, granularity: str) -> dict[str, object]:
        """Return the complete available monthly or yearly series for a category."""
        if granularity == "monthly":
            rows = self._rows(
                """SELECT to_char(t.transaction_date, 'YYYY-MM') AS period, SUM(t.amount) AS amount
                    FROM transactions t LEFT JOIN categories c ON c.id = t.category_id
                    WHERE t.user_id = :user_id AND t.direction = 'debit'
                      AND COALESCE(c.parent_id, c.id) = :category_id
                      AND t.transaction_kind NOT IN ('transfer', 'credit_card_payment')
                    GROUP BY period ORDER BY period""",
                {"user_id": self.user_id, "category_id": category_id},
            )
            amounts = {str(row["period"]): Decimal(str(row["amount"])) for row in rows}
            points = []
            if rows:
                start = date.fromisoformat(f"{rows[0]['period']}-01")
                end = date.fromisoformat(f"{rows[-1]['period']}-01")
                point = start
                while point <= end:
                    key = point.strftime("%Y-%m")
                    points.append({"period": key, "amount": amounts.get(key, Decimal("0"))})
                    point = date(point.year + (point.month == 12), 1 if point.month == 12 else point.month + 1, 1)
        else:
            rows = self._rows(
                """SELECT to_char(t.transaction_date, 'YYYY') AS period, SUM(t.amount) AS amount
                    FROM transactions t LEFT JOIN categories c ON c.id = t.category_id
                    WHERE t.user_id = :user_id AND t.direction = 'debit'
                      AND COALESCE(c.parent_id, c.id) = :category_id
                      AND t.transaction_kind NOT IN ('transfer', 'credit_card_payment')
                    GROUP BY period ORDER BY period""",
                {"user_id": self.user_id, "category_id": category_id},
            )
            points = [{"period": str(row["period"]), "amount": Decimal(str(row["amount"]))} for row in rows]
        return {"category_id": category_id, "granularity": granularity, "points": points}

    def monthly_insights(self, month: str) -> dict[str, object]:
        """Return deterministic, evidence-linked month insights.

        This intentionally contains facts and calculations only. Presentation or
        an LLM may explain these results later, but cannot create new anomalies.
        """
        year, month_number = (int(part) for part in month.split("-"))
        month_start = date(year, month_number, 1)
        month_end = date(year, month_number, calendar.monthrange(year, month_number)[1])
        previous_end = month_start - timedelta(days=1)
        previous_start = date(previous_end.year, previous_end.month, 1)
        rows = self._rows(
            """SELECT t.id, t.transaction_date, t.narration, t.merchant_normalized, t.amount,
                       COALESCE(c.name, 'Uncategorized') AS category
                FROM transactions t LEFT JOIN categories c ON c.id = t.category_id
                WHERE t.user_id = :user_id AND t.direction = 'debit'
                  AND t.transaction_kind NOT IN ('transfer', 'credit_card_payment')
                  AND t.transaction_date BETWEEN :start AND :end
                ORDER BY t.transaction_date""",
            {"user_id": self.user_id, "start": month_start, "end": month_end},
        )
        previous_categories = self._rows(
            """SELECT COALESCE(c.name, 'Uncategorized') AS category, SUM(t.amount) AS amount
                FROM transactions t LEFT JOIN categories c ON c.id = t.category_id
                WHERE t.user_id = :user_id AND t.direction = 'debit'
                  AND t.transaction_kind NOT IN ('transfer', 'credit_card_payment')
                  AND t.transaction_date BETWEEN :start AND :end
                GROUP BY c.name""",
            {"user_id": self.user_id, "start": previous_start, "end": previous_end},
        )
        if not rows:
            return {"month": month, "expense": Decimal("0"), "forecast": None, "anomalies": []}
        expense = sum((Decimal(str(row["amount"])) for row in rows), Decimal("0"))
        typical_amount = Decimal(str(median([Decimal(str(row["amount"])) for row in rows])))
        anomaly_floor = max(Decimal("5000"), typical_amount * Decimal("3"))
        anomalies: list[dict[str, object]] = []
        for row in rows:
            amount = Decimal(str(row["amount"]))
            if amount >= anomaly_floor:
                anomalies.append({"kind": "large_transaction", "title": "Unusually large transaction", "amount": amount,
                                  "transaction_id": row["id"], "transaction_date": row["transaction_date"],
                                  "merchant": row["merchant_normalized"] or row["narration"], "category": row["category"],
                                  "reason": "Amount is at least three times the typical transaction amount this month."})
        current_categories: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        for row in rows:
            current_categories[str(row["category"])] += Decimal(str(row["amount"]))
        previous_by_category = {str(row["category"]): Decimal(str(row["amount"])) for row in previous_categories}
        for category, amount in current_categories.items():
            previous = previous_by_category.get(category, Decimal("0"))
            if previous >= Decimal("1000") and amount >= previous * Decimal("1.5"):
                support = [row["id"] for row in rows if row["category"] == category]
                anomalies.append({"kind": "category_spike", "title": f"{category} spending increased", "amount": amount,
                                  "transaction_ids": support, "category": category,
                                  "reason": f"{category} is {((amount / previous) - 1) * 100:.0f}% higher than the prior month."})
        elapsed_days = max(1, min((max(row["transaction_date"] for row in rows) - month_start).days + 1, month_end.day))
        forecast = (expense / Decimal(elapsed_days) * Decimal(month_end.day)).quantize(Decimal("0.01"))
        return {"month": month, "expense": expense, "forecast": {"projected_expense": forecast, "days_observed": elapsed_days, "days_in_month": month_end.day}, "anomalies": anomalies[:10]}

    def detect_recurring_payments(self) -> dict[str, int]:
        """Persist only high-signal recurring debit patterns for user review.

        The detector is intentionally deterministic: an account/merchant pair needs
        three observations, a recognised cadence, and a bounded amount variation.
        """
        rows = self._rows(
            """SELECT t.financial_account_id, t.transaction_date, t.amount, t.narration,
                       t.merchant_normalized, t.category_id
                FROM transactions t
                WHERE t.user_id = :user_id AND t.direction = 'debit'
                  AND t.transaction_kind NOT IN ('transfer', 'credit_card_payment')
                ORDER BY t.financial_account_id, t.transaction_date"""
        )
        grouped: dict[tuple[UUID, str], list[dict[str, object]]] = defaultdict(list)
        for row in rows:
            merchant_key = _normalize_merchant(str(row["merchant_normalized"] or row["narration"]))
            if merchant_key:
                grouped[(row["financial_account_id"], merchant_key)].append(row)

        detections: list[dict[str, object]] = []
        for (account_id, merchant_key), entries in grouped.items():
            if len(entries) < 3:
                continue
            intervals = [(entries[index]["transaction_date"] - entries[index - 1]["transaction_date"]).days for index in range(1, len(entries))]
            cadence = _recurrence_cadence(intervals)
            if not cadence:
                continue
            cadence_name, cadence_days = cadence
            amounts = [Decimal(str(entry["amount"])) for entry in entries]
            typical_amount = Decimal(str(median(amounts)))
            tolerance = max(Decimal("50"), typical_amount * Decimal("0.10"))
            if any(abs(amount - typical_amount) > tolerance for amount in amounts):
                continue
            last = entries[-1]
            confidence = min(Decimal("0.97"), Decimal("0.72") + Decimal("0.05") * min(len(entries), 5))
            display_name = str(last["merchant_normalized"] or last["narration"]).strip()[:160]
            detections.append({
                "id": uuid4(), "user_id": self.user_id, "account_id": account_id,
                "merchant_key": merchant_key, "display_name": display_name,
                "category_id": last["category_id"], "cadence": cadence_name,
                "cadence_days": cadence_days, "typical_amount": typical_amount,
                "tolerance": tolerance, "occurrence_count": len(entries),
                "first_observed": entries[0]["transaction_date"], "last_observed": last["transaction_date"],
                "next_expected": last["transaction_date"] + timedelta(days=cadence_days), "confidence": confidence,
            })
        with Session(self.engine) as session, session.begin():
            for item in detections:
                session.execute(text("""INSERT INTO recurring_payment_detections
                    (id, user_id, financial_account_id, merchant_key, display_name, category_id, cadence, cadence_days,
                     typical_amount, amount_tolerance, occurrence_count, first_observed_on, last_observed_on,
                     next_expected_on, confidence)
                    VALUES (:id, :user_id, :account_id, :merchant_key, :display_name, :category_id, :cadence,
                            :cadence_days, :typical_amount, :tolerance, :occurrence_count, :first_observed,
                            :last_observed, :next_expected, :confidence)
                    ON CONFLICT (user_id, financial_account_id, merchant_key, cadence) DO UPDATE SET
                        display_name = EXCLUDED.display_name, category_id = EXCLUDED.category_id,
                        cadence_days = EXCLUDED.cadence_days, typical_amount = EXCLUDED.typical_amount,
                        amount_tolerance = EXCLUDED.amount_tolerance, occurrence_count = EXCLUDED.occurrence_count,
                        first_observed_on = EXCLUDED.first_observed_on, last_observed_on = EXCLUDED.last_observed_on,
                        next_expected_on = EXCLUDED.next_expected_on, confidence = EXCLUDED.confidence,
                        updated_at = now()"""), item)
        return {"detected": len(detections)}

    def list_recurring_payments(self, state: str | None = None) -> list[dict[str, object]]:
        parameters: dict[str, object] = {"user_id": self.user_id}
        query = """SELECT r.*, a.display_name AS account_name, c.name AS category
                   FROM recurring_payment_detections r
                   JOIN financial_accounts a ON a.id = r.financial_account_id
                   LEFT JOIN categories c ON c.id = r.category_id
                   WHERE r.user_id = :user_id"""
        if state:
            query += " AND r.state = :state"
            parameters["state"] = state
        rows = self._rows(query + " ORDER BY r.next_expected_on, r.display_name", parameters)
        subscription_hints = {
            "netflix", "spotify", "youtube", "icloud", "adobe", "microsoft",
            "prime", "membership", "openai", "chatgpt", "canva", "notion",
        }
        cadence_factor = {
            "weekly": Decimal("52") / Decimal("12"),
            "monthly": Decimal("1"),
            "quarterly": Decimal("1") / Decimal("3"),
            "yearly": Decimal("1") / Decimal("12"),
        }
        return [
            {
                **row,
                "kind": "subscription"
                if "subscription" in str(row["category"] or "").lower()
                or any(hint in str(row["merchant_key"]) for hint in subscription_hints)
                else "recurring",
                "monthly_equivalent": (
                    Decimal(str(row["typical_amount"])) * cadence_factor[str(row["cadence"])]
                ).quantize(Decimal("0.01")),
            }
            for row in rows
        ]

    def review_recurring_payment(self, detection_id: UUID, state: str) -> dict[str, object]:
        if state not in {"detected", "confirmed", "dismissed"}:
            raise LedgerError("Recurring payment state is invalid")
        with Session(self.engine) as session, session.begin():
            result = session.execute(text("""UPDATE recurring_payment_detections
                SET state = :state, updated_at = now() WHERE id = :id AND user_id = :user_id"""),
                {"state": state, "id": detection_id, "user_id": self.user_id})
            if result.rowcount != 1:
                raise LedgerError("Recurring payment detection was not found")
        return self._one("SELECT * FROM recurring_payment_detections WHERE id = :id AND user_id = :user_id", {"id": detection_id, "user_id": self.user_id})

    def update_recurring_payment(self, detection_id: UUID, payload: dict[str, object]) -> dict[str, object]:
        fields: list[str] = []
        parameters: dict[str, object] = {"id": detection_id, "user_id": self.user_id}
        if "display_name" in payload:
            fields.append("display_name = :display_name")
            parameters["display_name"] = _required_text(payload, "display_name")
        if "typical_amount" in payload:
            fields.append("typical_amount = :typical_amount")
            parameters["typical_amount"] = _positive_decimal(payload["typical_amount"], "typical_amount")
        if "next_expected_on" in payload:
            fields.append("next_expected_on = :next_expected_on")
            try:
                parameters["next_expected_on"] = date.fromisoformat(str(payload["next_expected_on"]))
            except ValueError as error:
                raise LedgerError("next_expected_on must use YYYY-MM-DD format") from error
        if not fields:
            raise LedgerError("No recurring-payment changes were provided")
        with Session(self.engine) as session, session.begin():
            result = session.execute(
                text(
                    f"""UPDATE recurring_payment_detections
                        SET {", ".join(fields)}, updated_at = now()
                        WHERE id = :id AND user_id = :user_id"""
                ),
                parameters,
            )
            if result.rowcount != 1:
                raise LedgerError("Recurring payment detection was not found")
        return self._one(
            "SELECT * FROM recurring_payment_detections WHERE id = :id AND user_id = :user_id",
            {"id": detection_id, "user_id": self.user_id},
        )

    def list_card_statements(self) -> list[dict[str, object]]:
        return self._rows(
            """SELECT s.id, s.financial_account_id, a.display_name AS account_name,
                      a.institution_code, s.period_start, s.period_end,
                      s.statement_amount, s.minimum_due, s.due_date,
                      s.total_limit, s.available_limit, s.state,
                      COALESCE(p.status, 'unpaid') AS payment_status,
                      COALESCE(p.paid_amount, 0) AS paid_amount, p.paid_at
               FROM statements s
               JOIN financial_accounts a ON a.id = s.financial_account_id
               LEFT JOIN card_statement_payments p ON p.statement_id = s.id
               WHERE s.user_id = :user_id AND a.account_type = 'credit_card'
                 AND s.state IN ('confirmed', 'reconciled')
               ORDER BY s.due_date DESC NULLS LAST, s.period_end DESC NULLS LAST"""
        )

    def update_card_statement_payment(self, statement_id: UUID, payload: dict[str, object]) -> dict[str, object]:
        status = _required_choice(payload, "status", {"unpaid", "partial", "paid"})
        paid_amount = Decimal("0")
        if payload.get("paid_amount") not in (None, ""):
            try:
                paid_amount = Decimal(str(payload["paid_amount"]))
            except InvalidOperation as error:
                raise LedgerError("paid_amount must be a valid amount") from error
            if paid_amount < 0:
                raise LedgerError("paid_amount cannot be negative")
        self._one(
            """SELECT s.id FROM statements s JOIN financial_accounts a
               ON a.id = s.financial_account_id
               WHERE s.id = :id AND s.user_id = :user_id
                 AND a.account_type = 'credit_card'""",
            {"id": statement_id, "user_id": self.user_id},
        )
        with Session(self.engine) as session, session.begin():
            session.execute(
                text(
                    """INSERT INTO card_statement_payments
                       (id, user_id, statement_id, status, paid_amount, paid_at)
                       VALUES (:id, :user_id, :statement_id, :status, :paid_amount,
                               CASE WHEN :status = 'paid' THEN now() ELSE NULL END)
                       ON CONFLICT (statement_id) DO UPDATE
                       SET status = EXCLUDED.status, paid_amount = EXCLUDED.paid_amount,
                           paid_at = CASE WHEN EXCLUDED.status = 'paid' THEN now() ELSE NULL END,
                           updated_at = now()"""
                ),
                {
                    "id": uuid4(),
                    "user_id": self.user_id,
                    "statement_id": statement_id,
                    "status": status,
                    "paid_amount": paid_amount,
                },
            )
        return self._one(
            """SELECT statement_id, status, paid_amount, paid_at
               FROM card_statement_payments
               WHERE statement_id = :id AND user_id = :user_id""",
            {"id": statement_id, "user_id": self.user_id},
        )

    def generate_card_reminders(self, today: date | None = None) -> dict[str, int]:
        current = today or date.today()
        statements = self._rows(
            """SELECT s.id, a.display_name AS account_name, s.statement_amount,
                      s.minimum_due, s.due_date, COALESCE(p.status, 'unpaid') AS payment_status
               FROM statements s JOIN financial_accounts a ON a.id = s.financial_account_id
               LEFT JOIN card_statement_payments p ON p.statement_id = s.id
               WHERE s.user_id = :user_id AND a.account_type = 'credit_card'
                 AND s.due_date IS NOT NULL AND s.state IN ('confirmed', 'reconciled')
                 AND COALESCE(p.status, 'unpaid') <> 'paid'
                 AND s.due_date <= :deadline""",
            {"user_id": self.user_id, "deadline": current + timedelta(days=7)},
        )
        created = 0
        with Session(self.engine) as session, session.begin():
            for statement in statements:
                overdue = statement["due_date"] < current
                kind = "card_payment_overdue" if overdue else "card_payment_upcoming"
                result = session.execute(
                    text(
                        """INSERT INTO notifications
                           (id, user_id, notification_kind, deduplication_key,
                            title, body, due_at)
                           VALUES (:id, :user_id, :kind, :key, :title, :body, :due_at)
                           ON CONFLICT (user_id, notification_kind, deduplication_key)
                           DO NOTHING"""
                    ),
                    {
                        "id": uuid4(),
                        "user_id": self.user_id,
                        "kind": kind,
                        "key": str(statement["id"]),
                        "title": f"{statement['account_name']} payment {'overdue' if overdue else 'due soon'}",
                        "body": (
                            f"Statement amount {statement['statement_amount'] or 0}; "
                            f"minimum due {statement['minimum_due'] or 0}."
                        ),
                        "due_at": statement["due_date"],
                    },
                )
                created += result.rowcount
        return {"created": created, "eligible": len(statements)}

    def list_notifications(self, state: str | None = None) -> list[dict[str, object]]:
        query = """SELECT id, notification_kind, title, body, state, due_at, created_at
                   FROM notifications WHERE user_id = :user_id"""
        parameters: dict[str, object] = {"user_id": self.user_id}
        if state:
            query += " AND state = :state"
            parameters["state"] = state
        return self._rows(query + " ORDER BY due_at NULLS LAST, created_at DESC", parameters)

    def update_notification(self, notification_id: UUID, state: str) -> dict[str, object]:
        if state not in {"unread", "read", "dismissed"}:
            raise LedgerError("Notification state is invalid")
        with Session(self.engine) as session, session.begin():
            result = session.execute(
                text(
                    """UPDATE notifications SET state = :state, updated_at = now()
                       WHERE id = :id AND user_id = :user_id"""
                ),
                {"id": notification_id, "user_id": self.user_id, "state": state},
            )
            if result.rowcount != 1:
                raise LedgerError("Notification was not found")
        return self._one(
            "SELECT * FROM notifications WHERE id = :id AND user_id = :user_id",
            {"id": notification_id, "user_id": self.user_id},
        )

    def balance_summary(self) -> dict[str, object]:
        rows = self._rows(
            """SELECT a.id, a.display_name, a.institution_code, a.account_type, a.currency,
            COALESCE(SUM(CASE WHEN t.direction = 'credit' THEN t.amount ELSE -t.amount END), 0) AS signed_total
            FROM financial_accounts a LEFT JOIN transactions t ON t.financial_account_id = a.id
            WHERE a.user_id = :user_id AND a.status = 'active'
            GROUP BY a.id, a.display_name, a.institution_code, a.account_type, a.currency ORDER BY a.display_name"""
        )
        cash = Decimal("0")
        liability = Decimal("0")
        accounts: list[dict[str, object]] = []
        for row in rows:
            balance = row["signed_total"] if row["account_type"] == "bank_account" else -row["signed_total"]
            accounts.append({**row, "balance": balance})
            if row["account_type"] == "bank_account":
                cash += balance
            else:
                liability += balance
        return {"cash_balance": cash, "credit_card_outstanding": liability, "net_worth": cash - liability, "accounts": accounts}

    def _require_account(self, account_id: UUID) -> None:
        self._one("SELECT id FROM financial_accounts WHERE id = :id AND user_id = :user_id AND status = 'active'", {"id": account_id, "user_id": self.user_id})

    def _rows(self, query: str, parameters: dict[str, object] | None = None) -> list[dict[str, object]]:
        with Session(self.engine) as session:
            return [dict(row) for row in session.execute(text(query), parameters or {"user_id": self.user_id}).mappings()]

    def _one(self, query: str, parameters: dict[str, object]) -> dict[str, object]:
        with Session(self.engine) as session:
            row = session.execute(text(query), parameters).mappings().one_or_none()
            if row is None:
                raise LedgerError("Requested resource was not found")
            return dict(row)


def _parse_tabular_upload(filename: str, content: bytes) -> list[dict[str, object]]:
    return _parse_tabular_upload_with_mapping(filename, content, None)


def _recurrence_cadence(intervals: list[int]) -> tuple[str, int] | None:
    """Return a recognised cadence only when all intervals are close to it."""
    if not intervals:
        return None
    candidate = round(float(median(intervals)))
    for name, target, allowed_variance in (
        ("weekly", 7, 2), ("monthly", 30, 6), ("quarterly", 91, 10), ("yearly", 365, 24),
    ):
        if abs(candidate - target) <= allowed_variance and all(abs(interval - target) <= allowed_variance for interval in intervals):
            return name, target
    return None


def _reporting_period_bounds(period: str, today: date | None = None) -> tuple[date | None, date | None]:
    if period not in REPORTING_PERIODS:
        raise LedgerError("Unsupported reporting period")
    if period == "all_time":
        return None, None
    current = today or date.today()
    current_month = date(current.year, current.month, 1)
    next_month = _shift_month(current_month, 1)
    if period == "this_month":
        return current_month, next_month
    if period == "last_month":
        return _shift_month(current_month, -1), current_month
    if period == "last_3_months":
        return _shift_month(current_month, -2), next_month
    if period == "last_6_months":
        return _shift_month(current_month, -5), next_month
    return date(current.year, 1, 1), date(current.year + 1, 1, 1)


def _shift_month(value: date, offset: int) -> date:
    month_index = value.year * 12 + value.month - 1 + offset
    return date(month_index // 12, month_index % 12 + 1, 1)


def _apply_period_filter(
    query: str,
    parameters: dict[str, object],
    column: str,
    period: str,
) -> tuple[str, dict[str, object]]:
    start, end = _reporting_period_bounds(period)
    if start is not None:
        query += f" AND {column} >= :period_start"
        parameters["period_start"] = start
    if end is not None:
        query += f" AND {column} < :period_end"
        parameters["period_end"] = end
    return query, parameters


def _month_start_end(month: str) -> tuple[date, date]:
    try:
        start = date.fromisoformat(f"{month}-01")
    except ValueError as error:
        raise LedgerError("Month must use YYYY-MM format") from error
    return start, _shift_month(start, 1)


def _positive_decimal(value: object, field: str) -> Decimal:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, TypeError) as error:
        raise LedgerError(f"{field} must be a valid amount") from error
    if not amount.is_finite() or amount <= 0:
        raise LedgerError(f"{field} must be greater than zero")
    return amount


def inspect_tabular_upload(filename: str, content: bytes) -> dict[str, object]:
    records = _tabular_records(filename, content)
    headers = list(records[0]) if records else []
    normalized_headers = {header.strip().lower(): header for header in headers}
    suggestions = {
        field: normalized_headers[candidate]
        for field, candidates in _MAPPING_CANDIDATES.items()
        for candidate in candidates
        if candidate in normalized_headers
    }
    return {"headers": headers, "suggested_mapping": suggestions, "sample_row_count": len(records)}


_MAPPING_CANDIDATES = {
    "transaction_date": ("transaction date", "date", "txn date"),
    "posted_date": ("value date", "value dt", "posted date"),
    "narration": ("transaction remarks", "narration", "description", "particulars"),
    "provider_reference": ("reference no.", "ref no", "chq./ref.no.", "utr"),
    "debit": ("withdrawal amount", "withdrawal amt.", "debit"),
    "credit": ("deposit amount", "deposit amt.", "credit"),
}


def _parse_tabular_upload_with_mapping(
    filename: str, content: bytes, column_mapping: dict[str, str] | None
) -> list[dict[str, object]]:
    normalized_rows, errors = _parse_tabular_upload_with_issues(filename, content, column_mapping)
    if errors:
        messages = [f"row {error['ordinal']}: {error['message']}" for error in errors]
        raise LedgerError("Statement validation failed — " + "; ".join(messages[:10]))
    return normalized_rows


def _parse_tabular_upload_with_issues(
    filename: str, content: bytes, column_mapping: dict[str, str] | None
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    records = _tabular_records(filename, content)
    if column_mapping:
        required = {"transaction_date", "narration", "debit", "credit"}
        missing = required - set(column_mapping)
        if missing:
            raise LedgerError(f"Column mapping is missing: {', '.join(sorted(missing))}")
        records = [_apply_column_mapping(record, column_mapping) for record in records]
    normalized_rows: list[dict[str, object]] = []
    errors: list[dict[str, object]] = []
    for ordinal, record in enumerate(records, start=2):
        try:
            normalized_rows.append(_normalize_row(record))
        except LedgerError as error:
            errors.append({"ordinal": ordinal, "message": str(error)})
    return normalized_rows, errors


def _tabular_records(filename: str, content: bytes) -> list[dict[str, object]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        if b"\x00" in content:
            raise LedgerError("CSV contains unsupported binary data")
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError as error:
            raise LedgerError("CSV must be UTF-8 encoded") from error
        return list(csv.DictReader(io.StringIO(decoded)))
    elif suffix == ".xlsx":
        if not content.startswith(b"PK"):
            raise LedgerError("XLSX file signature is invalid")
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        if len(workbook.sheetnames) > 10:
            raise LedgerError("XLSX workbook has too many sheets")
        worksheet = workbook.active
        values = list(worksheet.iter_rows(values_only=True))
        if len(values) > 100_001:
            raise LedgerError("XLSX worksheet has too many rows")
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        return [dict(zip(headers, row, strict=True)) for row in values[1:] if any(row)]
    else:
        raise LedgerError("Only CSV and XLSX statement imports are supported")


def _apply_column_mapping(record: dict[str, object], mapping: dict[str, str]) -> dict[str, object]:
    available = {str(key): value for key, value in record.items()}
    missing = [header for header in mapping.values() if header not in available]
    if missing:
        raise LedgerError(f"Mapped statement columns were not found: {', '.join(sorted(missing))}")
    return {
        "Transaction Date": available[mapping["transaction_date"]],
        "Value Date": available[mapping["posted_date"]] if mapping.get("posted_date") else None,
        "Transaction Remarks": available[mapping["narration"]],
        "Withdrawal Amount": available[mapping["debit"]],
        "Deposit Amount": available[mapping["credit"]],
        "Reference No.": available[mapping["provider_reference"]] if mapping.get("provider_reference") else None,
    }


def _normalize_row(raw: dict[str, object]) -> dict[str, object]:
    normalized = {str(key).strip().lower(): value for key, value in raw.items()}
    date_value = _first(normalized, "transaction date", "date", "txn date")
    narration = _first(normalized, "transaction remarks", "narration", "description", "particulars")
    reference = _first(normalized, "reference no.", "ref no", "chq./ref.no.", "utr")
    withdrawal = _money(_first(normalized, "withdrawal amount", "withdrawal amt.", "debit"))
    deposit = _money(_first(normalized, "deposit amount", "deposit amt.", "credit"))
    if (withdrawal is None) == (deposit is None):
        raise LedgerError("Each row must contain exactly one debit or credit amount")
    return {
        "transaction_date": _date(date_value),
        "posted_date": _date(_first(normalized, "value date", "value dt", "posted date"), optional=True),
        "narration": _text(narration, "narration"),
        "amount": withdrawal or deposit,
        "currency": "INR",
        "direction": "debit" if withdrawal is not None else "credit",
        "provider_reference": str(reference).strip() if reference not in (None, "") else None,
        "raw_columns": json.dumps({key: str(value) for key, value in raw.items()}, sort_keys=True),
    }


def _first(values: dict[str, object], *names: str) -> object | None:
    return next((values[name] for name in names if values.get(name) not in (None, "")), None)


def _money(value: object | None) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        amount = Decimal(str(value).replace(",", "").replace("₹", "").strip())
    except InvalidOperation as exc:
        raise LedgerError("A statement amount is invalid") from exc
    if amount <= 0:
        raise LedgerError("A statement amount must be positive")
    return amount


def _date(value: object | None, *, optional: bool = False) -> date | None:
    if value in (None, "") and optional:
        return None
    if isinstance(value, date):
        return value
    for pattern in ("%d/%m/%Y", "%d-%b-%Y", "%Y-%m-%d"):
        try:
            from datetime import datetime
            return datetime.strptime(str(value), pattern).date()
        except ValueError:
            continue
    raise LedgerError("A statement date is invalid")


def _transaction_kind(narration: str) -> str:
    upper = narration.upper()
    if any(marker in upper for marker in ("SALARY", "INTEREST", "DIVIDEND")):
        return "income"
    if any(marker in upper for marker in ("CARD PAYMENT", "CC PAYMENT", "CREDIT CARD PAYMENT", "CARD BILL", "BILL PAYMENT")):
        return "credit_card_payment"
    if any(marker in upper for marker in ("NEFT", "IMPS", "RTGS", "TRANSFER")):
        return "transfer"
    return "unknown"


def _normalize_merchant(value: str) -> str:
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value.lower()).split())


def _required_text(payload: dict[str, object], key: str) -> str:
    return _text(payload.get(key), key)


def _optional_text(payload: dict[str, object], key: str) -> str | None:
    value = payload.get(key)
    return str(value).strip() if value not in (None, "") else None


def _text(value: object | None, field: str) -> str:
    if value is None or not str(value).strip():
        raise LedgerError(f"{field} is required")
    return str(value).strip()


def _required_choice(payload: dict[str, object], key: str, allowed: set[str]) -> str:
    value = _required_text(payload, key)
    if value not in allowed:
        raise LedgerError(f"{key} is invalid")
    return value


def _safe_filename(value: str) -> str:
    return Path(value).name[:255]


def _encode_transaction_cursor(transaction_date: object, transaction_id: object) -> str:
    value = f"{transaction_date}|{transaction_id}".encode("ascii")
    return base64.urlsafe_b64encode(value).decode("ascii").rstrip("=")


def _decode_transaction_cursor(cursor: str) -> tuple[date, UUID]:
    try:
        padded = cursor + "=" * (-len(cursor) % 4)
        value = base64.urlsafe_b64decode(padded.encode("ascii")).decode("ascii")
        date_value, id_value = value.split("|", maxsplit=1)
        return date.fromisoformat(date_value), UUID(id_value)
    except (ValueError, UnicodeDecodeError, binascii.Error) as error:
        raise LedgerError("Transaction cursor is invalid") from error
